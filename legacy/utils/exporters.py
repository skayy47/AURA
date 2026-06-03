"""Multi-format export utilities for AURA — each function returns bytes."""
from __future__ import annotations

from io import BytesIO

import pandas as pd


def to_csv(df: pd.DataFrame) -> bytes:
    """Serialise *df* to UTF-8 CSV bytes."""
    buf = BytesIO()
    df.to_csv(buf, index=False)
    return buf.getvalue()


def to_excel(df: pd.DataFrame, sheet_name: str = "AURA") -> bytes:
    """Serialise *df* to xlsx bytes with styled header row (bold, blue background)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # Style header row
        header_fill = PatternFill(
            start_color="1D4ED8", end_color="1D4ED8", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit column widths (cap at 40)
        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df[col].empty else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                int(max_len) + 2, 40
            )

    return buf.getvalue()


def to_json(df: pd.DataFrame, orient: str = "records") -> bytes:
    """Serialise *df* to JSON bytes."""
    return df.to_json(orient=orient, indent=2).encode("utf-8")


def to_parquet(df: pd.DataFrame) -> bytes:
    """Serialise *df* to Parquet bytes."""
    buf = BytesIO()
    df.to_parquet(buf, index=False)
    return buf.getvalue()
